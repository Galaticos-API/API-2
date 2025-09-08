import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.text import Text
from openpyxl.styles import Font, Alignment

def criar_burndown_chart_excel(nome_arquivo, total_de_tarefas):
    """
    Cria uma planilha Excel com um gráfico de Burndown Chart.

    Args:
        nome_arquivo (str): O nome do arquivo .xlsx a ser criado.
        total_de_tarefas (int): O número de tarefas na sprint.
    """
    # Cria uma nova pasta de trabalho e seleciona a planilha ativa
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Burndown Chart"

    # --- Preenche os cabeçalhos da tabela ---
    headers = ["Total de horas"] + [f"Dia {i}" for i in range(1, 21)]
    ws.append([""] + headers)

    # Adiciona as linhas das tarefas
    for i in range(1, total_de_tarefas + 1):
        ws.append([f"Atividade {i}"] + [""] * 21)

    # --- Preenche a linha "Restante" e "Estimado" ---
    ws.append(["Restante"] + [""] * 21)
    ws.append(["Estimado"] + [""] * 21)

    # --- Estiliza os cabeçalhos e a primeira coluna ---
    for row in ws.iter_rows(min_row=1, max_row=total_de_tarefas + 3, min_col=1, max_col=22):
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Adiciona dados de exemplo ---
    # Coluna "Total de horas"
    horas_totais = [20, 10, 15, 10, 8, 12, 15, 8, 12, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]
    for i, horas in enumerate(horas_totais[:total_de_tarefas], start=2):
        ws[f'B{i}'] = horas

    # Célula de exemplo para a primeira tarefa no Dia 1
    ws['C2'] = 10
    # Célula de exemplo para a segunda tarefa no Dia 1
    ws['C3'] = 5
    # Célula de exemplo para a terceira tarefa no Dia 2
    ws['D4'] = 2

    # --- Cálculos para as linhas "Restante" e "Estimado" ---
    total_horas = sum(horas_totais[:total_de_tarefas])
    ws['B' + str(total_de_tarefas + 2)] = total_horas # Restante no Dia 0
    ws['B' + str(total_de_tarefas + 3)] = total_horas # Estimado no Dia 0

    # Lógica para o restante
    for dia in range(1, 21):
        col_letra = openpyxl.utils.get_column_letter(dia + 2)
        soma_dia_anterior = f'{openpyxl.utils.get_column_letter(dia+1)}{total_de_tarefas + 2}'
        soma_atividades_dia = f'SUM({col_letra}2:{col_letra}{total_de_tarefas + 1})'
        ws[col_letra + str(total_de_tarefas + 2)] = f'={soma_dia_anterior}-{soma_atividades_dia}'

    # Lógica para a linha estimada
    decaimento_diario = total_horas / 20
    for dia in range(1, 21):
        col_letra = openpyxl.utils.get_column_letter(dia + 2)
        ws[col_letra + str(total_de_tarefas + 3)] = total_horas - (decaimento_diario * dia)

   # --- Criação do Gráfico de Burndown ---
    chart = LineChart()
    chart.title = "Gráfico Burndown"
    chart.style = 10
    chart.y_axis.title = "Total de horas"
    chart.x_axis.title = "Dias"
    chart.y_axis.crossAx = 500
    chart.x_axis = openpyxl.chart.axis.DateAxis(crossAx=100)
    chart.x_axis.number_format = 'dd-mm-yyyy'


    # Adiciona a série "Restante"
    data_restante = Reference(ws, min_col=2, min_row=total_de_tarefas + 2, max_col=22, max_row=total_de_tarefas + 2)
    series_restante = openpyxl.chart.Series(data_restante, title="Restante")
    chart.series.append(series_restante)

    # Adiciona a série "Estimado"
    data_estimado = Reference(ws, min_col=2, min_row=total_de_tarefas + 3, max_col=22, max_row=total_de_tarefas + 3)
    series_estimado = openpyxl.chart.Series(data_estimado, title="Estimado")
    chart.series.append(series_estimado)

    # Configura o eixo X
    dates = Reference(ws, min_col=2, min_row=1, max_col=22)
    chart.set_categories(dates)

    # Adiciona o gráfico à planilha
    ws.add_chart(chart, "A" + str(total_de_tarefas + 5))

    # Salva o arquivo
    wb.save(nome_arquivo)
    print(f"Planilha '{nome_arquivo}' criada com sucesso!")

# --- Exemplo de uso ---
criar_burndown_chart_excel("burndown_sprint.xlsx", 20)